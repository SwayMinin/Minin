import matplotlib.pyplot as plt
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Side, Border, Font
from jinja2 import Environment, FileSystemLoader
import pdfkit

from input_connect import InputConnect


class Report:
    """Класс для создания отчетов со статистикой

    Attributes:
        default_font_size (int): Размер шрифта по умолчанию
        small_font_size (int): Размер маленького шрифта
        default_column_width (int): Ширина колонок по умолчанию
    """
    default_font_size = 8
    small_font_size = 6
    default_column_width = 2

    @staticmethod
    def generate_pdf(inp_connect):
        """Создает pdf файл со статистикой

        Args:
            inp_connect (InputConnect): Класс со статистикой
        """
        job_name = inp_connect.job_name
        Report.generate_image(inp_connect)
        Report.generate_excel(inp_connect)
        wb = load_workbook(filename='report.xlsx')
        years_rows = [[cell.value for cell in row] for row in wb.active.iter_rows()]
        cities_rows = [[cell.value for cell in row] for row in wb[wb.sheetnames[1]].iter_rows()]
        half = len(cities_rows[0]) // 2
        cities_salary_rows = []
        cities_share_rows = []
        for row in cities_rows:
            cities_salary_rows.append(row[:half])
            cities_share_rows.append(row[half+1:])
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render({'job_name': job_name,
                                        'years_headers': years_rows.pop(0), 'years_rows': years_rows,
                                        'cities_salary_headers': cities_salary_rows.pop(0),
                                        'cities_salary_rows': cities_salary_rows,
                                        'cities_share_headers': cities_share_rows.pop(0),
                                        'cities_share_rows': cities_share_rows})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    @staticmethod
    def generate_image(inp_connect: InputConnect):
        """Создает изображение с графиками статистики

        Args:
            inp_connect (InputConnect): Класс со статистикой
        """
        year_to_salary = inp_connect.year_to_salary
        year_to_vacancies_count = inp_connect.year_to_vacancies_count
        job_name = inp_connect.job_name
        job_year_to_salary = inp_connect.job_year_to_salary
        job_year_to_vacancies_count = inp_connect.job_year_to_vacancies_count
        city_to_salary = inp_connect.city_to_salary
        city_to_vacancies_share = inp_connect.city_to_vacancies_share

        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2)
        Report.__plot_comparison(ax1, "Уровень зарплат по годам",
                                 "cредняя з/п", f"з/п {job_name.lower()}",
                                 year_to_salary, job_year_to_salary)
        Report.__plot_comparison(ax2, "Количество вакансий по годам",
                                 "Количество вакансий", f"Количество вакансий\n{job_name.lower()}",
                                 year_to_vacancies_count, job_year_to_vacancies_count, )
        Report.__plot_cities_salary(ax3, "Уровень зарплат по городам", city_to_salary)
        Report.__plot_pie(ax4, "Доля вакансий по городам", city_to_vacancies_share)
        plt.tight_layout()
        plt.savefig('graph.png', dpi=300, bbox_inches='tight')

    @staticmethod
    def generate_excel(inp_connect: InputConnect):
        """Создает Excel таблицу со статистикой

        Args:
            inp_connect (InputConnect): Класс со статистикой
        """
        year_to_salary = inp_connect.year_to_salary
        year_to_vacancies_count = inp_connect.year_to_vacancies_count
        job_name = inp_connect.job_name
        job_year_to_salary = inp_connect.job_year_to_salary
        job_year_to_vacancies_count = inp_connect.job_year_to_vacancies_count
        city_to_salary = inp_connect.city_to_salary
        city_to_vacancies_share = inp_connect.city_to_vacancies_share

        wb = openpyxl.Workbook()
        # Первая страница
        ws1 = wb.active
        ws1.title = "Статистика по годам"
        Report.__make_headlines(ws1, ["Год", "Средняя зарплата", f"Средняя зарплата - {job_name}",
                                      "Количество вакансий", f"Количество вакансий - {job_name}"])
        for year in year_to_salary:
            ws1.append([year, year_to_salary[year], job_year_to_salary[year],
                        year_to_vacancies_count[year], job_year_to_vacancies_count[year]])
        Report.__auto_columns_width(ws1)
        # Вторая страница
        ws2 = wb.create_sheet()
        ws2.title = "Статистика по городам"
        Report.__make_headlines(ws2, ["Город", "Уровень зарплат", "Город", "Доля вакансий"])
        city_to_vacancies_share = {k: str(round(v * 100, 2)) + '%' for k, v in city_to_vacancies_share.items()}
        for city_salary, city_share in zip(city_to_salary, city_to_vacancies_share):
            ws2.append([city_salary, city_to_salary[city_salary],
                        city_share, city_to_vacancies_share[city_share]])
        for cell in ws2['D']:
            cell.number_format = '0.00%'
        ws2.insert_cols(3)
        ws2.column_dimensions['C'].width = Report.default_column_width
        Report.__auto_columns_width(ws2)
        # Сохранить
        wb.save('report.xlsx')

    @staticmethod
    def __make_headlines(worksheet, headlines):
        """Создает заголовки в листе таблицы

        Args:
            worksheet: Лист таблицы
            headlines (list): Список заголовков
        """
        worksheet.append(headlines)
        for cell in worksheet['1:1']:
            cell.font = Font(bold=True)
            column = worksheet.column_dimensions[cell.column_letter]
            column.width = len(str(cell.value)) + Report.default_column_width

    @staticmethod
    def __auto_columns_width(worksheet):
        """Задает ширину всем столбцам листа таблицы

        Args:
            worksheet: Лист таблицы
        """
        regular = Side(border_style="thin", color="000000")
        box = Border(top=regular, bottom=regular, left=regular, right=regular)
        for col in worksheet.iter_cols():
            for cell in col:
                if not cell.value:
                    continue
                cell.border = box
                column = worksheet.column_dimensions[cell.column_letter]
                column.width = max(column.width, len(str(cell.value)) + Report.default_column_width)

    @staticmethod
    def __plot_comparison(ax, title, label, label_compare, data, data_compare):
        """Задает на графике сравнение данных

        Args:
            ax: График
            title (str): Название графика
            label (str): Название данных
            label_compare (str): Название данных для сравнения
            data (dict): Данные
            data_compare (dict): Данные для сравнения
        """
        ax.set_title(title)
        bar_width = 0.4
        ax.bar([key - bar_width / 2 for key in data.keys()], data.values(), bar_width, label=label)
        ax.bar([key + bar_width / 2 for key in data_compare.keys()], data_compare.values(), bar_width, label=label_compare)
        ax.set_xticks(range(2007, 2023))
        ax.grid(axis='y')
        ax.legend(fontsize=Report.default_font_size)
        ax.tick_params('both', labelsize=Report.default_font_size)
        ax.tick_params('x', labelrotation=90)

    @staticmethod
    def __plot_cities_salary(ax, title, city_salary_data):
        """Задает на графике статистику зарплат по городам

        Args:
            ax: График
            title (str): Название графика
            city_salary_data (dict): Данные с зарплатой городов
        """
        ax.set_title(title)
        cities = [city.replace(' ', '\n').replace('-', '-\n') for city in city_salary_data.keys()]
        ax.barh(cities, city_salary_data.values())
        ax.invert_yaxis()
        ax.grid(axis='x')
        ax.tick_params('x', labelsize=Report.default_font_size)
        ax.tick_params('y', labelsize=Report.small_font_size)

    @staticmethod
    def __plot_pie(ax, title, city_share_data):
        """Задает круговую диаграмму со статистикой доли городов

        Args:
            title (str): Название графика
            city_share_data (dict): Данные с долей городов
        """
        ax.set_title(title)
        pie_dict = city_share_data.copy()
        pie_dict["Другие"] = 1 - sum(city_share_data.values())
        ax.pie(pie_dict.values(), labels=pie_dict.keys(), textprops={'fontsize': Report.small_font_size})
